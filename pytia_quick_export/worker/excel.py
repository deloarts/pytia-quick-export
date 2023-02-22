"""
    Export submodule. Holds utility functions for handling data exports.
"""
from dataclasses import asdict
from pathlib import Path
from typing import Literal

from helper.translators import translate_source, translate_type
from models.data import DataModel, DatumModel
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pytia.log import log
from resources import resource


def export_excel(
    path: Path,
    selected_project: str,
    data: DataModel,
    source: Literal["made", "bought"],
) -> None:
    """
    Exports the EXCEL file containing all the information of the document.
    The EXCEL file will have two rows, the header and the data row.
    For configuration see the 'excel.json' resource file.

    Args:
        path (Path): The path into which to save the EXCEL (xlsx) file.
        selected_project (str): The project number (from the UI).
        data (DataModel): The documents data to write.
        source (str): The source of the document (`made` or `bought`).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = selected_project

    _write_data(worksheet=ws, data=data)
    _style_worksheet(worksheet=ws)

    wb.save(str(path))
    log.info(f"Saved excel document to {str(path)!r}.")


def _write_data(
    worksheet: Worksheet,
    data: DataModel,
) -> None:
    """
    Saves the documents data to the EXCEL worksheet.

    Args:
        worksheet (Worksheet): The EXCEL worksheet.
        data (DataModel): The documents data to write.
    """
    for datum in data.data:
        _write_header(worksheet=worksheet, datum=datum)
        _write_datum(worksheet=worksheet, datum=datum)


def _write_header(worksheet: Worksheet, datum: DatumModel) -> None:
    """
    Writes a single header item to the EXCEL worksheet.

    Args:
        worksheet (Worksheet): The EXCEL worksheet.
        datum (DatumModel): The datum model from which to create the header.
    """
    if resource.excel.header_row is not None:
        header_cell = worksheet.cell(resource.excel.header_row + 1, datum.index + 1)
        cell_value = datum.name
        if isinstance(header_cell, Cell):
            header_cell.value = cell_value
            log.info(f"Wrote header {cell_value!r} to worksheet.")


def _write_datum(worksheet: Worksheet, datum: DatumModel) -> None:
    """
    Writes a single datum to the EXCEL worksheet.

    Args:
        worksheet (Worksheet): The EXCEL worksheet.
        datum (DatumModel): The datum model from which to write the data.
    """
    if datum.value is not None:
        datum_cell = worksheet.cell(resource.excel.data_row + 1, datum.index + 1)
        cell_value = datum.value
        if isinstance(datum_cell, Cell):
            datum_cell.value = cell_value
            log.info(f"Wrote value {cell_value!r} to worksheet.")


def _style_worksheet(worksheet: Worksheet) -> None:
    """Styles the worksheet as stated in the excel.json config file."""
    for column_cells in worksheet.columns:
        # Set format, font and size
        for index, cell in enumerate(column_cells):
            if index > resource.excel.data_row - 1:
                color = (
                    resource.excel.data_color_1
                    if index % 2 == 0
                    else resource.excel.data_color_2
                )
                bg_color = (
                    resource.excel.data_bg_color_1
                    if index % 2 == 0
                    else resource.excel.data_bg_color_2
                )
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
            else:
                color = None
            cell.number_format = "@"
            cell.font = Font(
                name=resource.excel.font, size=resource.excel.size, color=color
            )
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Set font and height for the header row
        if isinstance(resource.excel.header_row, int):
            worksheet.row_dimensions[resource.excel.header_row + 1].height = 20  # type: ignore
            column_cells[resource.excel.header_row].font = Font(
                name=resource.excel.font,
                size=resource.excel.size,
                bold=True,
                color=resource.excel.header_color,
            )
            column_cells[resource.excel.header_row].fill = PatternFill(
                start_color=resource.excel.header_bg_color,
                end_color=resource.excel.header_bg_color,
                fill_type="solid",
            )
            column_cells[resource.excel.header_row].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Set cell width
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = (  # type: ignore
            length if length > 2 else 2
        )

    log.info(f"Styled worksheet {worksheet.title!r}.")
